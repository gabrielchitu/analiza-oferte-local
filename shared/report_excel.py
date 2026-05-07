import io
from datetime import date
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

GRAY_FILL = PatternFill("solid", fgColor="C0C0C0")
SUBHDR_FILL = PatternFill("solid", fgColor="D9D9D9")
RED_FILL = PatternFill("solid", fgColor="FFCCCC")
ORANGE_FILL = PatternFill("solid", fgColor="FFB347")
BOLD = Font(bold=True)
RED_BOLD = Font(bold=True, color="CC0000")
RED_FONT = Font(color="CC0000")
THIN = Side(style="thin")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _observatie(neconf: dict) -> str:
    tip = neconf.get("tip", "")
    if tip == "ARTICOL_LIPSA":
        return "LIPSĂ ARTICOL"
    if tip == "ARTICOL_EXTRA":
        return "ARTICOL SUPLIMENTAR ÎN OFERTĂ"
    if tip == "UM_DIFERIT":
        ref_um = neconf.get("ref_um", "")
        off_um = neconf.get("oferta_um", "")
        return f"UNITATE DE MĂSURĂ DIFERITĂ\nRef: {ref_um}  |  Ofertă: {off_um}"
    if tip == "DIFERENTA_CAMP":
        camp = neconf.get("camp", "")
        labels = {
            "cantitate": "CANTITATE DIFERITĂ",
            "pret_material": "PREȚ MATERIAL DIFERIT",
            "pret_manopera": "PREȚ MANOPERĂ DIFERIT",
            "pret_utilaj": "PREȚ UTILAJ DIFERIT",
            "pret_transport": "PREȚ TRANSPORT DIFERIT",
            "val_material": "VALOARE MATERIAL DIFERITĂ",
            "val_manopera": "VALOARE MANOPERĂ DIFERITĂ",
            "val_utilaj": "VALOARE UTILAJ DIFERITĂ",
            "val_transport": "VALOARE TRANSPORT DIFERITĂ",
        }
        label = labels.get(camp, f"DIFERENȚĂ {camp.upper()}")
        return f"{label}\nRef: {neconf.get('ref', '')}  |  Ofertă: {neconf.get('oferta', '')}"
    if tip == "EROARE_ARITMETICA":
        camp = neconf.get("camp", "")
        return (f"EROARE ARITMETICĂ: {camp}\n"
                f"Declarat: {neconf.get('declarat', '')}  Calculat: {neconf.get('calculat', '')}")
    if tip == "COD_SIMILAR":
        ref_cod = neconf.get("ref_cod", "")
        oferta_cod = neconf.get("oferta_cod", "")
        motiv = neconf.get("motiv_similaritate", "")
        return (f"COD SIMILAR — posibilă eroare OCR sau variație\n"
                f"Ref: {ref_cod}  |  Ofertat: {oferta_cod}\n{motiv}")
    return tip


def _should_exclude_field(field_name: str, comparison_mode: str) -> bool:
    """Return True if field should be excluded for fara_pret mode."""
    if comparison_mode == "fara_pret":
        return field_name.startswith("pret_") or field_name.startswith("val_")
    return False


def _write_row(ws, row_idx: int, nr: int, nc: dict, comparison_mode: str = "cu_pret"):
    tip = nc.get("tip", "")

    # Col A: Nr. crt.
    ws.cell(row_idx, 1, nr).border = BORDER

    # Col B: Cerinta (referinta)
    if tip != "ARTICOL_EXTRA":
        ref_cod = nc.get("ref_cod", "")
        ref_um = nc.get("ref_um", "")
        ref_cant = nc.get("ref_cantitate", "")
        ref_den = nc.get("ref_denumire", "")
        b_text = f"{ref_cod}  {ref_um}  {ref_cant}\n{ref_den}" if ref_den else f"{ref_cod}  {ref_um}  {ref_cant}"
    else:
        b_text = "—"
    c_b = ws.cell(row_idx, 2, b_text)
    c_b.alignment = Alignment(wrap_text=True, vertical="top")
    c_b.border = BORDER

    # Col C: Ce a ofertat
    if tip != "ARTICOL_LIPSA":
        off_cod = nc.get("oferta_cod", "")
        off_um = nc.get("oferta_um", "")
        off_cant = nc.get("oferta_cantitate", "")
        off_den = nc.get("oferta_denumire", "")
        c_text = f"{off_cod}  {off_um}  {off_cant}\n{off_den}" if off_den else f"{off_cod}  {off_um}  {off_cant}"
    else:
        c_text = "—"
    c_c = ws.cell(row_idx, 3, c_text)
    c_c.alignment = Alignment(wrap_text=True, vertical="top")
    c_c.border = BORDER

    # Col D: Observatii
    obs = _observatie(nc)
    c_d = ws.cell(row_idx, 4, obs)
    c_d.font = RED_BOLD
    c_d.alignment = Alignment(wrap_text=True, vertical="top")
    c_d.border = BORDER

    # Col E: Suspect (cu_pret mode only)
    if comparison_mode == "cu_pret":
        suspect_text = ""
        if nc.get("suspect", False):
            motiv = nc.get("motiv_suspiciune", "")
            suspect_text = f"⚠️ {motiv}" if motiv else "⚠️"
        c_e = ws.cell(row_idx, 5, suspect_text)
        c_e.alignment = Alignment(wrap_text=True, vertical="top")
        c_e.border = BORDER

    # Highlight row if there are differences
    if tip in ("ARTICOL_LIPSA", "ARTICOL_EXTRA", "UM_DIFERIT", "DIFERENTA_CAMP", "EROARE_ARITMETICA"):
        cols_to_highlight = 5 if comparison_mode == "cu_pret" else 4
        for col in range(1, cols_to_highlight + 1):
            ws.cell(row_idx, col).fill = RED_FILL
    elif tip == "COD_SIMILAR":
        cols_to_highlight = 5 if comparison_mode == "cu_pret" else 4
        for col in range(1, cols_to_highlight + 1):
            ws.cell(row_idx, col).fill = ORANGE_FILL


def generate_excel(session: dict, comparatii: list, comparison_mode: str = "cu_pret") -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    sumar_data = []
    has_systemic_alert = False
    systemic_alert_msg = ""

    for comp in comparatii:
        nr = comp.get("oferta_nr", "?")
        source = comp.get("source_file", "")
        ofertant = comp.get("ofertant", "")
        ws = wb.create_sheet(title=f"Oferta_{nr}")

        # Column widths — adjust for cu_pret mode with Suspect column
        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 40
        ws.column_dimensions["C"].width = 40
        ws.column_dimensions["D"].width = 35
        if comparison_mode == "cu_pret":
            ws.column_dimensions["E"].width = 20

        # Meta rows
        r = 1
        ws.cell(r, 1, "Beneficiar:").font = BOLD
        ws.cell(r, 2, session.get("client_name", ""))
        r += 1
        ws.cell(r, 1, "Obiect investiție:").font = BOLD
        ws.cell(r, 2, session.get("obiect_investitii", ""))
        r += 1
        ws.cell(r, 1, "Data:").font = BOLD
        ws.cell(r, 2, str(date.today()))
        r += 1
        ws.cell(r, 1, "Oferta:").font = BOLD
        ws.cell(r, 2, f"Oferta {nr} — {source}")
        if ofertant:
            r += 1
            ws.cell(r, 1, "Ofertant:").font = BOLD
            ws.cell(r, 2, ofertant)
        r += 2

        # Header row
        col3_hdr = f"Ce a ofertat\n{ofertant}" if ofertant else "Ce a ofertat"
        headers = ["Nr.\ncrt.", "Cerința din caietul de sarcini", col3_hdr, "Observații"]
        if comparison_mode == "cu_pret":
            headers.append("Suspect")

        for col, hdr in enumerate(headers, 1):
            c = ws.cell(r, col, hdr)
            c.font = BOLD
            c.fill = GRAY_FILL
            c.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
            c.border = BORDER
        ws.row_dimensions[r].height = 30
        r += 1

        neconformitati = comp.get("neconformitati", [])
        if not neconformitati:
            ws.cell(r, 1, "Nicio neconcordanță detectată.")
            sumar_data.append((f"Oferta {nr}", 0))
            continue

        # Group by deviz
        by_deviz = defaultdict(list)
        for nc in neconformitati:
            key = nc.get("deviz_denumire", "NECUNOSCUT")
            by_deviz[key].append(nc)

        nr_crt = 1
        for deviz_den, neconfs in by_deviz.items():
            # Deviz group header
            merge_to_col = 5 if comparison_mode == "cu_pret" else 4
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=merge_to_col)
            c = ws.cell(r, 1, f"Categoria de lucrari: {deviz_den}")
            c.font = BOLD
            c.fill = SUBHDR_FILL
            c.alignment = Alignment(wrap_text=True, vertical="center")
            c.border = BORDER
            ws.row_dimensions[r].height = 20
            r += 1

            for nc in neconfs:
                _write_row(ws, r, nr_crt, nc, comparison_mode)
                ws.row_dimensions[r].height = 40
                nr_crt += 1
                r += 1

        sumar_data.append((f"Oferta {nr}", len(neconformitati)))

        # Check for systemic alert
        verif_sondaj = comp.get("verificare_sondaj", {})
        if verif_sondaj.get("sistemic", False):
            has_systemic_alert = True
            suspect_ratio = verif_sondaj.get("suspect_ratio", 0)
            systemic_alert_msg = f"⚠️ ALERTĂ SISTEMICĂ: {int(suspect_ratio * 100)}% din itemii verificați sunt suspecți"

    # Sumar sheet
    ws_sum = wb.create_sheet(title="Sumar", index=0)
    ws_sum.column_dimensions["A"].width = 20
    ws_sum.column_dimensions["B"].width = 25
    r = 1
    ws_sum.cell(r, 1, "Beneficiar:").font = BOLD
    ws_sum.cell(r, 2, session.get("client_name", ""))
    r += 1
    ws_sum.cell(r, 1, "Obiect investiție:").font = BOLD
    ws_sum.cell(r, 2, session.get("obiect_investitii", ""))
    r += 2

    # Add systemic alert if present
    if has_systemic_alert:
        c_alert = ws_sum.cell(r, 1, systemic_alert_msg)
        c_alert.font = Font(bold=True, color="CC0000", size=12)
        c_alert.fill = PatternFill("solid", fgColor="FFCCCC")
        c_alert.alignment = Alignment(wrap_text=True, vertical="center")
        ws_sum.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        ws_sum.row_dimensions[r].height = 25
        r += 2

    for col, hdr in enumerate(["Oferta", "Total neconcordanțe"], 1):
        c = ws_sum.cell(r, col, hdr)
        c.font = BOLD
        c.fill = GRAY_FILL
        c.border = BORDER
    r += 1
    for oferta, total in sumar_data:
        ws_sum.cell(r, 1, oferta).border = BORDER
        c = ws_sum.cell(r, 2, total)
        c.border = BORDER
        if total > 0:
            c.font = RED_BOLD
        r += 1

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
