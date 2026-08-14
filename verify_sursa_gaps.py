#!/usr/bin/env python3
"""Verificare gap-uri nr_crt în sursa_extracted_*.json vs raw DI JSON.

Usage:
    python3 verify_sursa_gaps.py --client "EuroProject" --json di_referinta
    python3 verify_sursa_gaps.py --client "EuroProject" --json di_oferta_1
    python3 verify_sursa_gaps.py --client "EuroProject"   # interactiv

Output:
    - Console: lista gap-uri cu clasificare (bug extractor / absent / OK)
    - Sheet "Erori" adăugat în XLSX-ul existent cu detalii gap-uri bug extractor
"""

import argparse
import json
import re
import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment

INPUT_BASE = Path("input_AO")
OUTPUT_BASE = Path("output_AO")

_NR_STANDALONE_RE = re.compile(r'^\d{1,3}$')
_TABLE_HEADER_LINE = '5 = 3 x 4'
_XLS_BOLD   = Font(bold=True)
_XLS_RED_BG = PatternFill('solid', fgColor='FFCCCC')
_XLS_WARN_BG= PatternFill('solid', fgColor='FFF2CC')
_XLS_HDR_BG = PatternFill('solid', fgColor='D9D9D9')
_XLS_CENTER = Alignment(horizontal='center', vertical='center')
_XLS_LEFT   = Alignment(horizontal='left',   vertical='center', wrap_text=True)


# ── raw DI helpers ──────────────────────────────────────────────────────────

def _load_di_pages(di_path: Path, f3_pages: set[int] | None = None) -> dict[int, list[str]]:
    """Raw lines per page, limited to the F3 pages the extractor actually read.

    The annex tables ('Lista materiale', 'Ore Manopera', 'Utilaje', 'Transport')
    are not F3 and carry their own numbering plus their own column-header rows
    ('4 = 2 X 3', '6 = 2 X 3 X 5'), which would otherwise surface as missing
    article numbers.
    """
    data = json.loads(di_path.read_text(encoding='utf-8'))
    result = {}
    for page in data.get('pages', []):
        pn = page.get('page_number', 0)
        if f3_pages is not None and pn not in f3_pages:
            continue
        lines = [ln.get('content', '').strip() for ln in page.get('lines', [])]
        result[pn] = lines
    return result


def _f3_page_numbers(client: str, json_stem: str) -> set[int] | None:
    """Page numbers classified as F3, from the classifier checkpoint."""
    ckpt = OUTPUT_BASE / client / f"{json_stem}_page_classes.json"
    if not ckpt.exists():
        return None
    data = json.loads(ckpt.read_text(encoding='utf-8'))
    pcs = data.get('page_classes', data) if isinstance(data, dict) else data
    return {pc.get('page_number') for pc in pcs if pc.get('is_f3')}


def _find_nr_in_pages(pages: dict[int, list[str]], nr: int) -> list[tuple[int, int, list[str]]]:
    """Return [(page_nr, line_idx, context_lines)] for standalone nr in raw pages.

    Lines above the '5 = 3 x 4' table header are skipped: the column numbers
    printed there ('0' '1' '2' '3' '4') are not article numbers.
    """
    hits = []
    nr_str = str(nr)
    for pn, lines in sorted(pages.items()):
        start = lines.index(_TABLE_HEADER_LINE) + 1 if _TABLE_HEADER_LINE in lines else 0
        for i, line in enumerate(lines):
            if i < start:
                continue
            if line.strip() == nr_str and _NR_STANDALONE_RE.match(line.strip()):
                context = lines[max(0, i - 2): i + 6]
                hits.append((pn, i, context))
    return hits


# ── extracted articles ──────────────────────────────────────────────────────

def _load_extracted(extracted_path: Path) -> list[dict]:
    data = json.loads(extracted_path.read_text(encoding='utf-8'))
    arts = []
    for deviz in data:
        for cap in deviz.get('capitole', []):
            for art in cap.get('articole', []):
                arts.append(art)
    return arts


def _main_nrs(arts: list[dict]) -> list[int]:
    nrs = []
    for a in arts:
        nr = a.get('nr_crt', '')
        try:
            nrs.append(int(nr))
        except (ValueError, TypeError):
            pass
    return sorted(set(nrs))


def _find_gaps(nrs: list[int]) -> list[int]:
    if len(nrs) < 2:
        return []
    full = set(range(nrs[0], nrs[-1] + 1))
    return sorted(full - set(nrs))


# ── XLSX sheet "Erori" ──────────────────────────────────────────────────────

def _write_errors_sheet(xlsx_path: Path, bug_gaps: list[dict], ok_gaps: list[int]) -> None:
    """Add/replace sheet 'Erori' in existing XLSX with gap findings."""
    wb = load_workbook(str(xlsx_path))

    # remove existing Erori sheet if present
    if 'Erori' in wb.sheetnames:
        del wb['Erori']

    ws = wb.create_sheet(title='Erori')

    headers = ['Nr.crt lipsă', 'Tip', 'Pagina raw', 'Linie idx', 'Context (±2 linii)']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = _XLS_BOLD
        cell.fill = _XLS_HDR_BG
        cell.alignment = _XLS_CENTER

    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 60

    row = 2
    for entry in bug_gaps:
        nr = entry['nr']
        for hit in entry['hits']:
            pn, idx, ctx = hit
            ctx_str = ' | '.join(ctx)
            ws.cell(row=row, column=1, value=nr).alignment = _XLS_CENTER
            ws.cell(row=row, column=2, value='BUG EXTRACTOR').fill = _XLS_RED_BG
            ws.cell(row=row, column=3, value=pn).alignment = _XLS_CENTER
            ws.cell(row=row, column=4, value=idx).alignment = _XLS_CENTER
            c = ws.cell(row=row, column=5, value=ctx_str)
            c.alignment = _XLS_LEFT
            row += 1

    for nr in ok_gaps:
        ws.cell(row=row, column=1, value=nr).alignment = _XLS_CENTER
        ws.cell(row=row, column=2, value='ABSENT (salt numerotare)').fill = _XLS_WARN_BG
        ws.cell(row=row, column=3, value='-').alignment = _XLS_CENTER
        ws.cell(row=row, column=4, value='-').alignment = _XLS_CENTER
        ws.cell(row=row, column=5, value='Nr. nu există în raw DI').alignment = _XLS_LEFT
        row += 1

    if row == 2:
        ws.cell(row=2, column=1, value='✓ Fără gap-uri').font = Font(color='008000', bold=True)

    wb.save(str(xlsx_path))


# ── core analysis ───────────────────────────────────────────────────────────

def analyze(client: str, json_stem: str) -> None:
    di_path = INPUT_BASE / client / f"{json_stem}.json"
    extracted_path = OUTPUT_BASE / client / f"sursa_extracted_{json_stem}.json"

    # find XLSX output — exclude lock files (~$...)
    stem_part = json_stem.replace('di_', '')
    acronym_glob = [f for f in (OUTPUT_BASE / client).glob(f"*{stem_part}*.xlsx")
                    if not f.name.startswith('~')]
    if not acronym_glob:
        acronym_glob = [f for f in (OUTPUT_BASE / client).glob("*.xlsx")
                        if not f.name.startswith('~')]
    xlsx_path = acronym_glob[0] if acronym_glob else None

    if not di_path.exists():
        print(f"[ERROR] {di_path} not found")
        sys.exit(1)
    if not extracted_path.exists():
        print(f"[ERROR] {extracted_path} not found — run gen_sursa_incarcare first")
        sys.exit(1)

    print(f"\nVerificare gap-uri: {client} / {json_stem}")
    print(f"  DI raw:    {di_path}")
    print(f"  Extras:    {extracted_path}")
    print(f"  XLSX:      {xlsx_path or 'negăsit'}")

    di_pages = _load_di_pages(di_path, _f3_page_numbers(client, json_stem))
    arts = _load_extracted(extracted_path)
    nrs = _main_nrs(arts)
    gaps = _find_gaps(nrs)

    print(f"\n  Articole extrase: {len(arts)}, nr range: {nrs[0] if nrs else '?'}-{nrs[-1] if nrs else '?'}")

    bug_gaps = []
    ok_gaps = []

    if gaps:
        print(f"\n  Gap-uri detectate: {gaps}")
        for nr in gaps:
            hits = _find_nr_in_pages(di_pages, nr)
            if hits:
                print(f"\n  ⚠ Nr {nr} — BUG EXTRACTOR (există în raw DI, pagina {hits[0][0]}):")
                for pn, idx, ctx in hits:
                    print(f"      p.{pn} linia {idx}: {ctx}")
                bug_gaps.append({'nr': nr, 'hits': hits})
            else:
                print(f"  ○ Nr {nr} — absent din raw DI (salt numerotare)")
                ok_gaps.append(nr)
    else:
        print("  ✓ Fără gap-uri în nr_crt.")

    # tail check: articole după max extras
    tail_missing = []
    if nrs:
        for nr in range(nrs[-1] + 1, nrs[-1] + 16):
            hits = _find_nr_in_pages(di_pages, nr)
            if hits:
                print(f"\n  ⚠ Nr {nr} — MISSING TAIL (există în raw DI, pagina {hits[0][0]}):")
                for pn, idx, ctx in hits:
                    print(f"      p.{pn}: {ctx}")
                tail_missing.append({'nr': nr, 'hits': hits})
            else:
                break

    total_bugs = len(bug_gaps) + len(tail_missing)
    print(f"\n  Sumar: {total_bugs} bug-uri extractor ({len(bug_gaps)} gap + {len(tail_missing)} tail), {len(ok_gaps)} salturi numerotare")

    if xlsx_path:
        _write_errors_sheet(xlsx_path, bug_gaps + tail_missing, ok_gaps)
        print(f"  → Sheet 'Erori' {'scris' if total_bugs or ok_gaps else 'actualizat (OK)'} în {xlsx_path.name}")
    else:
        print("  [WARN] XLSX negăsit — sheet 'Erori' nu a fost scris")


# ── CLI ─────────────────────────────────────────────────────────────────────

def _pick_client() -> str:
    from shared.client_config import ClientConfig
    clients = ClientConfig.detect_clients(INPUT_BASE)
    if not clients:
        print(f"No clients in {INPUT_BASE}/")
        sys.exit(1)
    print("\nClienti:")
    for i, c in enumerate(clients, 1):
        print(f"  {i}. {c}")
    choice = input("Client [numar]: ").strip()
    try:
        return clients[int(choice) - 1]
    except (ValueError, IndexError):
        print("Selectie invalida.")
        sys.exit(1)


def _pick_json(client: str) -> str:
    input_dir = INPUT_BASE / client
    files = sorted(input_dir.glob("di_*.json"))
    if not files:
        print(f"No di_*.json in {input_dir}")
        sys.exit(1)
    print(f"\nFisiere JSON:")
    for i, f in enumerate(files, 1):
        print(f"  {i}. {f.name}")
    choice = input("JSON [numar]: ").strip()
    try:
        return files[int(choice) - 1].stem
    except (ValueError, IndexError):
        print("Selectie invalida.")
        sys.exit(1)


def main() -> None:
    parser = argparse.ArgumentParser(description='Verificare gap-uri sursa incarcare')
    parser.add_argument('--client', help='Numele clientului')
    parser.add_argument('--json', dest='json_stem', metavar='JSON_STEM',
                        help='Stem JSON (ex: di_referinta, di_oferta_1)')
    args = parser.parse_args()

    client = args.client or _pick_client()
    json_stem = args.json_stem or _pick_json(client)
    if not json_stem.startswith('di_'):
        json_stem = f"di_{json_stem}"

    analyze(client, json_stem)


if __name__ == '__main__':
    main()
